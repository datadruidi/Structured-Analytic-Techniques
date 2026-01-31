#!/usr/bin/env python3
# -*- coding: utf-8 -*-

"""
Interactive Obsidian .md generator/updater from Excel (.xlsx).

Append-only updater:
- If a note exists, it is NOT rewritten.
- Only appends missing Parents/Children links and new Description/Source blocks.
- Preserves any manual edits anywhere in the note.
- If a note doesn't exist, creates it with a minimal template.

Features:
- Simple / Advanced mode
- Config save/load (obsidian_sat_config.json)
- Header preview
- Option to skip header row
- Dry-run summary + confirmation
"""

from __future__ import annotations

import json
import os
import re
import sys
from collections import defaultdict
from dataclasses import dataclass
from typing import Dict, List, Optional, Set, Tuple

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None

WIKILINK_RE = re.compile(r"\[\[([^\]|]+)(?:\|[^\]]+)?\]\]")
DEFAULT_CONFIG_FILENAME = "obsidian_sat_config.json"


# -------------------------
# Config
# -------------------------

@dataclass
class Config:
    input_path: str
    output_dir: str
    sheet_name: Optional[str]
    has_header_row: bool
    hierarchy_cols: List[str]       # e.g. ["A","B","C"]
    include_desc: bool
    desc_col: Optional[str]         # e.g. "F"
    source_col: Optional[str]       # e.g. "G"
    mode: str                       # "simple" or "advanced"

    def to_dict(self) -> dict:
        return {
            "input_path": self.input_path,
            "output_dir": self.output_dir,
            "sheet_name": self.sheet_name,
            "has_header_row": self.has_header_row,
            "hierarchy_cols": self.hierarchy_cols,
            "include_desc": self.include_desc,
            "desc_col": self.desc_col,
            "source_col": self.source_col,
            "mode": self.mode,
        }

    @staticmethod
    def from_dict(d: dict) -> "Config":
        return Config(
            input_path=str(d.get("input_path", "")).strip(),
            output_dir=str(d.get("output_dir", "")).strip(),
            sheet_name=(d.get("sheet_name") or None),
            has_header_row=bool(d.get("has_header_row", True)),
            hierarchy_cols=[str(x).strip().upper() for x in (d.get("hierarchy_cols", []) or [])],
            include_desc=bool(d.get("include_desc", True)),
            desc_col=(str(d.get("desc_col")).strip().upper() if d.get("desc_col") else None),
            source_col=(str(d.get("source_col")).strip().upper() if d.get("source_col") else None),
            mode=str(d.get("mode", "simple")).strip().lower() or "simple",
        )


# -------------------------
# Utility
# -------------------------

def sanitize_filename(name: str) -> str:
    """Filesystem-safe filename stem. Keeps Unicode, strips forbidden chars."""
    name = (name or "").strip()
    name = re.sub(r'[\\/:*?"<>|]', " - ", name)
    name = re.sub(r"\s+", " ", name).strip()
    name = name.rstrip(" .")
    return name or "Untitled"


def col_letter_to_index(col: str) -> int:
    """Excel column letters -> 0-based index (A->0, B->1, Z->25, AA->26)."""
    col = (col or "").strip().upper()
    if not re.fullmatch(r"[A-Z]+", col):
        raise ValueError(f"Invalid column '{col}'. Use e.g. A, B, C, AA.")
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def index_to_col_letter(idx0: int) -> str:
    """0-based index -> Excel column letters."""
    if idx0 < 0:
        raise ValueError("Negative index not allowed.")
    n = idx0 + 1
    letters = []
    while n:
        n, rem = divmod(n - 1, 26)
        letters.append(chr(ord("A") + rem))
    return "".join(reversed(letters))


def ask_path(prompt: str, must_exist: bool = True) -> str:
    while True:
        p = input(prompt).strip().strip('"')
        if not p:
            print("Please provide a path.", file=sys.stderr)
            continue
        if p == ".":
            p = os.getcwd()
        if must_exist and not os.path.exists(p):
            print(f"Path not found: {p}", file=sys.stderr)
            continue
        return p


def ask_yes_no(prompt: str, default: Optional[bool] = None) -> bool:
    suffix = ""
    if default is True:
        suffix = " [Y/n]"
    elif default is False:
        suffix = " [y/N]"

    while True:
        ans = input(prompt + suffix + ": ").strip().lower()
        if ans == "" and default is not None:
            return default
        if ans in ("y", "yes"):
            return True
        if ans in ("n", "no"):
            return False
        print("Answer Y/N.", file=sys.stderr)


def ask_choice(prompt: str, choices: List[str], default: Optional[str] = None) -> str:
    choices_l = [c.lower() for c in choices]
    while True:
        suffix = f" ({'/'.join(choices)})"
        if default:
            suffix += f" [default {default}]"
        ans = input(prompt + suffix + ": ").strip().lower()
        if ans == "" and default:
            ans = default.lower()
        if ans in choices_l:
            return ans
        print(f"Choose one of: {', '.join(choices)}", file=sys.stderr)


def ask_col(prompt: str) -> str:
    while True:
        ans = input(prompt).strip().upper()
        if re.fullmatch(r"[A-Z]+", ans):
            return ans
        print("Enter an Excel column letter, e.g. A, B, C, AA.", file=sys.stderr)


def load_config_if_present(config_path: str) -> Optional[Config]:
    if not os.path.exists(config_path):
        return None
    try:
        with open(config_path, "r", encoding="utf-8") as f:
            d = json.load(f)
        cfg = Config.from_dict(d)
        if not cfg.input_path or not cfg.output_dir or not cfg.hierarchy_cols:
            return None
        return cfg
    except Exception:
        return None


def save_config(config_path: str, cfg: Config) -> None:
    with open(config_path, "w", encoding="utf-8") as f:
        json.dump(cfg.to_dict(), f, indent=2, ensure_ascii=False)


# -------------------------
# Excel reading
# -------------------------

def open_workbook(xlsx_path: str, sheet_name: Optional[str]) -> Tuple[object, object]:
    if load_workbook is None:
        raise RuntimeError("openpyxl is not installed. Install it with: python -m pip install openpyxl")
    wb = load_workbook(xlsx_path, data_only=True)
    if sheet_name:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' not found. Available: {', '.join(wb.sheetnames)}")
        ws = wb[sheet_name]
    else:
        ws = wb.worksheets[0]
    return wb, ws


def preview_header_row(ws, max_preview_cols: int = 30) -> None:
    max_col = min(ws.max_column or 0, max_preview_cols)
    if max_col <= 0:
        print("No cells found to preview.")
        return
    print("\nHeader preview (row 1):")
    for i in range(max_col):
        v = ws.cell(row=1, column=i + 1).value
        s = "" if v is None else str(v).strip()
        print(f"  {index_to_col_letter(i)}: {s}")
    if ws.max_column and ws.max_column > max_preview_cols:
        print(f"  ... (preview limited to first {max_preview_cols} columns)")
    print("")


def validate_cols_exist(ws, cols: List[str]) -> None:
    if ws.max_column is None or ws.max_column <= 0:
        raise ValueError("Worksheet appears empty (max_column <= 0).")
    max_idx0 = ws.max_column - 1
    max_letter = index_to_col_letter(max_idx0)
    for c in cols:
        idx = col_letter_to_index(c)
        if idx > max_idx0:
            raise ValueError(f"Column '{c}' is outside sheet range. Max column is '{max_letter}'.")


def read_xlsx_rows_by_columns(
    ws,
    hierarchy_cols_letters: List[str],
    include_desc: bool,
    desc_col_letter: Optional[str],
    source_col_letter: Optional[str],
    start_row: int,
) -> List[dict]:
    hierarchy_idx0 = [col_letter_to_index(c) for c in hierarchy_cols_letters]
    desc_idx0 = col_letter_to_index(desc_col_letter) if (include_desc and desc_col_letter) else None
    src_idx0 = col_letter_to_index(source_col_letter) if (include_desc and source_col_letter) else None

    out: List[dict] = []
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        if row is None:
            continue

        parts: List[str] = []
        for ci in hierarchy_idx0:
            val = row[ci] if ci < len(row) else None
            sval = "" if val is None else str(val).strip()
            if sval:
                parts.append(sval)

        if not parts:
            continue

        description = ""
        source = ""
        if include_desc:
            if desc_idx0 is not None:
                v = row[desc_idx0] if desc_idx0 < len(row) else None
                description = "" if v is None else str(v).strip()
            if src_idx0 is not None:
                v = row[src_idx0] if src_idx0 < len(row) else None
                source = "" if v is None else str(v).strip()

        out.append({"parts": parts, "description": description, "source": source})

    return out


# -------------------------
# Graph building
# -------------------------

def build_title_graph_from_parts(rows: List[dict]) -> Tuple[
    Dict[str, Set[str]],
    Dict[str, Set[str]],
    Dict[str, List[Tuple[str, str]]],
    int,
    int,
]:
    parent_to_children: Dict[str, Set[str]] = defaultdict(set)
    child_to_parents: Dict[str, Set[str]] = defaultdict(set)
    leaf_info: Dict[str, List[Tuple[str, str]]] = defaultdict(list)

    edge_count = 0
    leaf_entry_count = 0

    for r in rows:
        parts: List[str] = r["parts"]
        if not parts:
            continue

        for i in range(len(parts) - 1):
            p = parts[i]
            c = parts[i + 1]
            if c not in parent_to_children[p]:
                edge_count += 1
            parent_to_children[p].add(c)
            child_to_parents[c].add(p)

        leaf_title = parts[-1]
        description = (r.get("description") or "").strip()
        source = (r.get("source") or "").strip()
        if description or source:
            leaf_info[leaf_title].append((description, source))
            leaf_entry_count += 1

        parent_to_children.setdefault(leaf_title, set())
        child_to_parents.setdefault(parts[0], set())

    return parent_to_children, child_to_parents, leaf_info, edge_count, leaf_entry_count


def dry_run_summary(
    parent_to_children: Dict[str, Set[str]],
    child_to_parents: Dict[str, Set[str]],
    leaf_info: Dict[str, List[Tuple[str, str]]],
    edge_count: int,
    leaf_entry_count: int,
) -> None:
    all_titles: Set[str] = set(parent_to_children.keys()) | set(child_to_parents.keys()) | set(leaf_info.keys())
    roots = sorted([t for t in all_titles if len(child_to_parents.get(t, set())) == 0], key=lambda x: x.lower())
    sample_titles = sorted(list(all_titles), key=lambda x: x.lower())[:15]

    print("\n=== Dry-run summary ===")
    print(f"Nodes (notes): {len(all_titles)}")
    print(f"Edges (parent->child links): {edge_count}")
    print(f"Leaf entries (description/source rows): {leaf_entry_count}")
    if roots:
        print(f"Root nodes (no parents): {', '.join(roots[:10])}" + (" ..." if len(roots) > 10 else ""))
    print("Sample notes:")
    for t in sample_titles:
        print(f"  - {t}")
    if len(all_titles) > len(sample_titles):
        print(f"  ... (showing first {len(sample_titles)} of {len(all_titles)})")
    print("")


# -------------------------
# Append-only Markdown updating
# -------------------------

def find_section_range(lines: List[str], header: str) -> Tuple[int, int]:
    """
    Find section '## <header>' range.
    Returns (start_idx, end_idx) where end_idx is first next '## ' or EOF.
    If not found -> (-1, -1)
    """
    hline = f"## {header}"
    start = -1
    for i, ln in enumerate(lines):
        if ln.strip() == hline:
            start = i
            break
    if start == -1:
        return -1, -1

    end = len(lines)
    for j in range(start + 1, len(lines)):
        if lines[j].startswith("## "):
            end = j
            break
    return start, end


def ensure_section(lines: List[str], header: str) -> Tuple[List[str], int, int]:
    """
    Ensure '## <header>' exists.
    If missing, append it at end (with a blank line before).
    Returns updated lines and section range (start,end).
    """
    s, e = find_section_range(lines, header)
    if s != -1:
        return lines, s, e

    # append missing section
    if lines and lines[-1].strip() != "":
        lines.append("")
    lines.append(f"## {header}")
    lines.append("")  # section body starts after this blank line
    return lines, len(lines) - 2, len(lines)


def extract_existing_wikilinks(lines: List[str], start: int, end: int) -> Set[str]:
    found: Set[str] = set()
    for ln in lines[start:end]:
        for m in WIKILINK_RE.finditer(ln):
            found.add(m.group(1).strip())
    return found


def normalize_leaf_pair(description: str, source: str) -> str:
    key = (description or "").strip() + "\n" + (source or "").strip()
    key = re.sub(r"\s+", " ", key).strip().lower()
    return key


def extract_existing_leaf_pairs(lines: List[str], start: int, end: int) -> Set[str]:
    """
    Looks for patterns:
      - **Description:** ...
        **Source:** ...
    and returns normalized keys for de-dup.
    """
    existing: Set[str] = set()
    desc = None
    src = None

    for ln in lines[start:end]:
        t = ln.strip()
        if t.startswith("- **Description:**"):
            desc = t[len("- **Description:**"):].strip()
            src = None
        elif desc is not None and t.startswith("**Source:**"):
            src = t[len("**Source:**"):].strip()
            existing.add(normalize_leaf_pair(desc, src))
            desc, src = None, None

    return existing


def append_missing_links(lines: List[str], section_header: str, new_links: Set[str]) -> List[str]:
    lines, s, e = ensure_section(lines, section_header)

    existing = extract_existing_wikilinks(lines, s, e)
    to_add = [x for x in sorted(new_links, key=lambda v: v.lower()) if x not in existing]

    if not to_add:
        return lines

    insert_at = e  # append at end of section
    # ensure there's a blank line after header if empty section
    # (we already add one when creating)
    for item in to_add:
        lines.insert(insert_at, f"- [[{item}]]")
        insert_at += 1

    return lines


def append_missing_leaf_blocks(lines: List[str], new_leaf_pairs: List[Tuple[str, str]]) -> List[str]:
    lines, s, e = ensure_section(lines, "Description and source")

    existing = extract_existing_leaf_pairs(lines, s, e)

    to_add_blocks: List[List[str]] = []
    for desc, src in new_leaf_pairs:
        key = normalize_leaf_pair(desc, src)
        if key and key not in existing:
            to_add_blocks.append([
                f"- **Description:** {desc}".rstrip(),
                f"  **Source:** {src}".rstrip(),
            ])
            existing.add(key)

    if not to_add_blocks:
        return lines

    insert_at = e
    for block in to_add_blocks:
        for ln in block:
            lines.insert(insert_at, ln)
            insert_at += 1

    return lines


def create_new_note_lines(title: str) -> List[str]:
    return [
        f"# {title}",
        "",
        "## Parents",
        "",
        "## Children",
        "",
        "## Description and source",
        "",
    ]


def update_note_file_append_only(
    file_path: str,
    title: str,
    add_parents: Set[str],
    add_children: Set[str],
    add_leaf_pairs: List[Tuple[str, str]],
) -> None:
    if os.path.exists(file_path):
        with open(file_path, "r", encoding="utf-8") as f:
            raw = f.read()
        lines = raw.splitlines()
        if not lines:
            lines = create_new_note_lines(title)
    else:
        lines = create_new_note_lines(title)

    # Append-only updates
    if add_parents:
        lines = append_missing_links(lines, "Parents", add_parents)
    if add_children:
        lines = append_missing_links(lines, "Children", add_children)
    if add_leaf_pairs:
        lines = append_missing_leaf_blocks(lines, add_leaf_pairs)

    # Ensure newline at end
    out = "\n".join(lines).rstrip() + "\n"
    with open(file_path, "w", encoding="utf-8", newline="\n") as f:
        f.write(out)


# -------------------------
# Interactive flow
# -------------------------

def interactive_config(script_dir: str) -> Config:
    print("\n=== Obsidian SAT - interactive generation ===\n")

    config_path = os.path.join(script_dir, DEFAULT_CONFIG_FILENAME)
    existing_cfg = load_config_if_present(config_path)

    if existing_cfg:
        use_prev = ask_yes_no(f"Found '{DEFAULT_CONFIG_FILENAME}'. Use it?", default=True)
        if use_prev:
            print("\nLoaded config. Press Enter to keep existing values.\n")

            ip = input(f"Input path (.xlsx) [{existing_cfg.input_path}]: ").strip().strip('"')
            if ip:
                existing_cfg.input_path = ip

            od = input(f"Output directory [{existing_cfg.output_dir}]: ").strip().strip('"')
            if od:
                existing_cfg.output_dir = (os.getcwd() if od == "." else od)

            sn = input(f"Sheet name (blank = first) [{existing_cfg.sheet_name or 'first'}]: ").strip()
            if sn:
                existing_cfg.sheet_name = sn

            return existing_cfg

    input_path = ask_path("Enter the input file path (.xlsx): ", must_exist=True)
    if not input_path.lower().endswith(".xlsx"):
        print("This script expects an Excel .xlsx file.", file=sys.stderr)
        return interactive_config(script_dir)

    output_dir = ask_path('Enter the output folder path ("." means current folder)', must_exist=False)
    os.makedirs(output_dir, exist_ok=True)

    sheet_name = input("Sheet name (leave blank to use the first sheet): ").strip()
    sheet_name = sheet_name if sheet_name else None

    mode = ask_choice("Mode", ["simple", "advanced"], default="simple")

    _, ws = open_workbook(input_path, sheet_name)
    preview_header_row(ws)

    has_header_row = ask_yes_no("Does row 1 contain headers (should it be skipped)?", default=True)

    print("Define hierarchy columns (e.g. A=top level, B=Level 1, ...).")

    hierarchy_cols: List[str] = []
    if mode == "simple":
        level_count_str = input("How many hierarchy levels TOTAL (including top level)? [default 5]: ").strip()
        level_count = 5
        if level_count_str:
            try:
                level_count = int(level_count_str)
                if level_count < 1:
                    raise ValueError()
            except Exception:
                print("Invalid number; using default 5.", file=sys.stderr)
                level_count = 5

        for i in range(level_count):
            label = "Top level" if i == 0 else f"Level {i}"
            col = ask_col(f"Which column contains {label} (e.g. {'A' if i == 0 else 'B'}): ")
            hierarchy_cols.append(col)

    else:
        top_col = ask_col("Which column contains the top-level concepts (e.g. A): ")
        hierarchy_cols.append(top_col)

        level1_col = ask_col("Which column contains Level 1 (e.g. B): ")
        hierarchy_cols.append(level1_col)

        level_n = 2
        while True:
            more = ask_yes_no(f"Add another level (next is Level {level_n})?", default=False)
            if not more:
                break
            col = ask_col(f"Which column contains Level {level_n} (e.g. C): ")
            hierarchy_cols.append(col)
            level_n += 1

    include_desc = ask_yes_no("Include descriptions and sources?", default=True)
    desc_col = None
    source_col = None
    if include_desc:
        desc_col = ask_col("Which column contains Description (e.g. F): ")
        source_col = ask_col("Which column contains Source (e.g. G): ")

    cfg = Config(
        input_path=input_path,
        output_dir=output_dir,
        sheet_name=sheet_name,
        has_header_row=has_header_row,
        hierarchy_cols=[c.upper() for c in hierarchy_cols],
        include_desc=include_desc,
        desc_col=(desc_col.upper() if desc_col else None),
        source_col=(source_col.upper() if source_col else None),
        mode=mode,
    )

    if ask_yes_no(f"Save this config to '{DEFAULT_CONFIG_FILENAME}' in the script folder?", default=True):
        save_config(config_path, cfg)
        print(f"Saved: {config_path}\n")

    return cfg


# -------------------------
# Main
# -------------------------

def main() -> int:
    script_dir = os.path.dirname(os.path.abspath(__file__))
    cfg = interactive_config(script_dir)

    if not os.path.exists(cfg.input_path):
        print(f"Input file not found: {cfg.input_path}", file=sys.stderr)
        return 2
    if not cfg.input_path.lower().endswith(".xlsx"):
        print("Input must be .xlsx", file=sys.stderr)
        return 2

    os.makedirs(cfg.output_dir, exist_ok=True)

    _, ws = open_workbook(cfg.input_path, cfg.sheet_name)

    cols_to_check = list(cfg.hierarchy_cols)
    if cfg.include_desc:
        if cfg.desc_col:
            cols_to_check.append(cfg.desc_col)
        if cfg.source_col:
            cols_to_check.append(cfg.source_col)

    try:
        validate_cols_exist(ws, cols_to_check)
    except Exception as e:
        print(f"Column validation error: {e}", file=sys.stderr)
        max_letter = index_to_col_letter((ws.max_column or 1) - 1)
        print(f"Tip: max sheet column is {max_letter}", file=sys.stderr)
        return 3

    start_row = 2 if cfg.has_header_row else 1

    rows = read_xlsx_rows_by_columns(
        ws=ws,
        hierarchy_cols_letters=cfg.hierarchy_cols,
        include_desc=cfg.include_desc,
        desc_col_letter=cfg.desc_col,
        source_col_letter=cfg.source_col,
        start_row=start_row,
    )

    if not rows:
        print("No rows produced. Check column choices and header skipping.", file=sys.stderr)
        return 4

    parent_to_children, child_to_parents, leaf_info, edge_count, leaf_entry_count = build_title_graph_from_parts(rows)

    all_titles: Set[str] = set(parent_to_children.keys()) | set(child_to_parents.keys()) | set(leaf_info.keys())
    if not all_titles:
        print("No nodes found. Check your data and column choices.", file=sys.stderr)
        return 5

    dry_run_summary(parent_to_children, child_to_parents, leaf_info, edge_count, leaf_entry_count)

    if not ask_yes_no("Proceed to write/update notes (append-only)?", default=True):
        print("Cancelled. No files written.")
        return 0

    # Append-only update per note
    for title in sorted(all_titles, key=lambda x: x.lower()):
        filename = sanitize_filename(title) + ".md"
        fpath = os.path.join(cfg.output_dir, filename)

        add_parents = child_to_parents.get(title, set())
        add_children = parent_to_children.get(title, set())
        add_leaf_pairs = leaf_info.get(title, [])

        update_note_file_append_only(
            file_path=fpath,
            title=title,
            add_parents=set(add_parents),
            add_children=set(add_children),
            add_leaf_pairs=list(add_leaf_pairs),
        )

    print(f"Done! Updated/created {len(all_titles)} notes in:\n{cfg.output_dir}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
